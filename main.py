from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os


# Create a new workbook object
wb= Workbook()


# Create a new xlsx file
filepath = 'data.xlsx'
    
if not os.path.exists(filepath):
    sheet = wb.active
    wb.save(filepath)
    
    
# load existing spreadsheet
wb=load_workbook('data.xlsx')


# Create a active worksheet
ws = wb.active

# Set a variable
name = ws["A2"].value
city =ws["B2"].value

# Print something from our Spreadsheet
print(f'{name}:{city}')

# Grab a whole columns
column_a = ws["A"]

# For loop
for cell in column_a:
    print(cell.value)
    
# Grab a range
range= ws["A2:B10"]


# For loop
for cell in range:
    for i in cell:
        print(i.value)
        
        
# Display sheets
sheets= wb.sheetnames 
print(sheets)      

# Get sheet
sheet1= wb['sheet1']

# Get max rows and columns
rows = sheet1.max_row
columns = sheet1.max_column

#Get Data 
for i in range(1,rows+1):
    for j in range(1,columns+1):
        print(sheet1.cell(i,j).value)
        
wb.save('data.xlsx')        

sheet1["A1"].value = "pass"
sheet1["A1"].fill=PatternFill("solid",fgColor="#55FF33")
sheet1["A2"].value = "fail"
sheet1["A2"].fill=PatternFill("solid",fgColor="#FF3346")